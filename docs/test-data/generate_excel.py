"""Generate test-data.xlsx from test-data.json using openpyxl."""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(SCRIPT_DIR, "test-data.json")
XLSX_PATH = os.path.join(SCRIPT_DIR, "test-data.xlsx")


# ── colour palette ──────────────────────────────────────────────────────────
TEAL_DARK   = "1A6B6E"
TEAL_MID    = "2D9B9F"
TEAL_LIGHT  = "D6F0F1"
SAGE        = "6BAE8C"
SAGE_LIGHT  = "E4F5EC"
AMBER       = "F59E0B"
AMBER_LIGHT = "FEF3C7"
RED_LIGHT   = "FEE2E2"
GREY_HEADER = "F1F5F9"
WHITE       = "FFFFFF"


def fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def bold_font(size=11, colour="000000") -> Font:
    return Font(bold=True, size=size, color=colour)


def header_font() -> Font:
    return Font(bold=True, size=10, color=WHITE)


def thin_border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_row(ws, headers: list[str], row=1,
                     bg=TEAL_DARK, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = header_font()
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        c.border = thin_border()


def stripe(ws, row: int, is_odd: bool):
    """Apply alternating row fill."""
    bg = WHITE if is_odd else GREY_HEADER
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        if c.fill.patternType != "solid" or c.fill.fgColor.rgb == "00000000":
            c.fill = fill(bg)


def set_col_widths(ws, widths: dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def write_cell(ws, row, col, value, bg=None, bold=False, wrap=False,
               align="left", num_format=None, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal=align, vertical="top",
                             wrap_text=wrap)
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    if bold or color != "000000":
        c.font = Font(bold=bold, color=color)
    if num_format:
        c.number_format = num_format
    return c


# ── Sheet 1: Products ───────────────────────────────────────────────────────
def sheet_products(wb, data):
    ws = wb.create_sheet("Products")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Name", "Category", "Unit", "Serving Amount",
        "Serving Label", "kcal", "Protein (g)", "Fat (g)", "Carbs (g)",
        "Diet Tags", "User Added?", "Owner ID", "This Week", "Next Week", "Notes"
    ]
    apply_header_row(ws, headers)

    widths = {1:8, 2:24, 3:14, 4:8, 5:14, 6:20, 7:7, 8:12, 9:10,
              10:10, 11:40, 12:12, 13:10, 14:11, 15:11, 16:35}
    set_col_widths(ws, widths)

    cat_colours = {
        "Dairy": "F0F9FF", "Fish": "EFF6FF", "Grains": "FFFBEB",
        "Produce": "F0FDF4", "Meat": "FFF1F2", "Legumes": "FAF5FF",
        "Nuts & Seeds": "FFF7ED", "Condiments": "F8FAFC"
    }

    for idx, p in enumerate(data["products"]):
        r = idx + 2
        row_bg = cat_colours.get(p["category"], WHITE)
        tags = ", ".join(p.get("dietTags", []))
        wf = p.get("weekFlags", {})
        note = p.get("_notes", "")
        vals = [
            p["id"], p["name"], p["category"], p["unit"],
            p["servingAmount"], p.get("servingLabel", ""),
            p["kcal"], p["protein"], p["fat"], p["carbs"],
            tags,
            "Yes" if p["isUserAdded"] else "No",
            p.get("userId") or "—",
            "✓" if wf.get("thisWeek") else "",
            "✓" if wf.get("nextWeek") else "",
            note
        ]
        for col, v in enumerate(vals, 1):
            bg = row_bg
            if col == 12 and v == "Yes":
                bg = SAGE_LIGHT
            write_cell(ws, r, col, v, bg=bg, wrap=(col in (11, 16)),
                       align="center" if col in (1,4,5,7,8,9,10,12,13,14,15) else "left")

    ws.row_dimensions[1].height = 30
    for r in range(2, len(data["products"]) + 2):
        ws.row_dimensions[r].height = 18


# ── Sheet 2: Recipes ─────────────────────────────────────────────────────────
def sheet_recipes(wb, data):
    ws = wb.create_sheet("Recipes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Name", "Category", "Servings",
        "kcal/Serving", "Protein (g)", "Fat (g)", "Carbs (g)",
        "Diet Tags", "Favourite?", "User Added?", "Owner ID",
        "This Week", "Next Week",
        "Ingredients (productId | amount unit)", "Notes"
    ]
    apply_header_row(ws, headers)

    widths = {1:8, 2:26, 3:15, 4:9, 5:13, 6:13, 7:10, 8:10,
              9:40, 10:11, 11:12, 12:10, 13:11, 14:11, 15:60, 16:40}
    set_col_widths(ws, widths)

    for idx, r in enumerate(data["recipes"]):
        row = idx + 2
        row_bg = SAGE_LIGHT if r["isUserAdded"] else WHITE
        tags = ", ".join(r.get("dietTags", []))
        wf = r.get("weekFlags", {})
        ings = " | ".join(
            f"{i['productId']} {i['amount']}{i['unit']}"
            for i in r.get("ingredients", [])
        )
        note = r.get("_notes", "")
        vals = [
            r["id"], r["name"], r["category"], r["servings"],
            r["kcalPerServing"], r["protein"], r["fat"], r["carbs"],
            tags,
            "Yes" if r.get("favorite") else "No",
            "Yes" if r["isUserAdded"] else "No",
            r.get("userId") or "—",
            "✓" if wf.get("thisWeek") else "",
            "✓" if wf.get("nextWeek") else "",
            ings, note
        ]
        for col, v in enumerate(vals, 1):
            write_cell(ws, row, col, v, bg=row_bg, wrap=(col in (9, 15, 16)),
                       align="center" if col in (1,4,5,6,7,8,10,11,12,13,14) else "left")

    ws.row_dimensions[1].height = 30
    for r in range(2, len(data["recipes"]) + 2):
        ws.row_dimensions[r].height = 45


# ── Sheet 3: Diets ────────────────────────────────────────────────────────────
def sheet_diets(wb, data):
    ws = wb.create_sheet("Diets")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Name", "Description",
        "Macro Guidance?", "Protein %", "Fat %", "Carbs %", "Notes"
    ]
    apply_header_row(ws, headers)
    set_col_widths(ws, {1:20, 2:28, 3:60, 4:16, 5:11, 6:9, 7:10, 8:40})

    for idx, d in enumerate(data["diets"]):
        r = idx + 2
        mg = d.get("macroGuidance") or {}
        has_macro = d.get("applicableMacroGuidance", False)
        row_bg = TEAL_LIGHT if has_macro else GREY_HEADER
        note = d.get("guidanceNote", "")
        vals = [
            d["id"], d["name"], d["description"],
            "Yes" if has_macro else "No",
            mg.get("proteinPct", "—"),
            mg.get("fatPct", "—"),
            mg.get("carbsPct", "—"),
            note
        ]
        for col, v in enumerate(vals, 1):
            write_cell(ws, r, col, v, bg=row_bg, wrap=(col in (3, 8)),
                       align="center" if col in (4, 5, 6, 7) else "left")

    ws.row_dimensions[1].height = 30
    for r in range(2, len(data["diets"]) + 2):
        ws.row_dimensions[r].height = 40


# ── Sheet 4: Users ────────────────────────────────────────────────────────────
def sheet_users(wb, data):
    ws = wb.create_sheet("Users")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "ID", "Email", "Password (test only)", "Role", "Test Role",
        "Language", "Unit System", "Gender", "Age", "Weight (kg)",
        "Active Diet", "Calorie Target", "Protein %", "Fat %", "Carbs %",
        "Calorie Corridor"
    ]
    apply_header_row(ws, headers)
    set_col_widths(ws, {1:8, 2:28, 3:20, 4:10, 5:22,
                        6:10, 7:14, 8:10, 9:6, 10:12,
                        11:22, 12:15, 13:11, 14:10, 15:10, 16:20})

    row_bgs = [TEAL_LIGHT, SAGE_LIGHT, AMBER_LIGHT]
    for idx, u in enumerate(data["users"]):
        r = idx + 2
        dp = u.get("dietPreferences", {})
        pr = u.get("profile", {})
        mac = dp.get("macros", {})
        vals = [
            u["id"], u["email"], u["password"], u.get("role", "User"),
            u.get("_testRole", ""),
            pr.get("language", ""), pr.get("unitSystem", ""),
            pr.get("gender") or "—", pr.get("age") or "—",
            pr.get("weightKg") or "—",
            dp.get("activeDiet") or "—",
            dp.get("calorieTarget") or "—",
            mac.get("proteinPct") or "—",
            mac.get("fatPct") or "—",
            mac.get("carbsPct") or "—",
            dp.get("calorieCorridorNote") or "—"
        ]
        bg = row_bgs[idx % len(row_bgs)]
        for col, v in enumerate(vals, 1):
            write_cell(ws, r, col, v, bg=bg,
                       align="center" if col in (1,4,6,7,8,9,10,12,13,14,15) else "left")

    ws.row_dimensions[1].height = 30
    for r in range(2, len(data["users"]) + 2):
        ws.row_dimensions[r].height = 22


# ── Sheet 5: Planner Seeds ────────────────────────────────────────────────────
def sheet_planner(wb, data):
    ws = wb.create_sheet("Planner Seeds")
    ws.sheet_view.showGridLines = False

    # -- Assignments section --
    ws["A1"] = "PLANNER ASSIGNMENTS (current week)"
    ws["A1"].font = bold_font(12, WHITE)
    ws["A1"].fill = fill(TEAL_DARK)
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["Assignment ID", "Item ID", "Item Name", "Day", "Meal Slot",
               "Servings", "kcal (est.)"]
    apply_header_row(ws, headers, row=2, bg=TEAL_MID)

    seeds = data["plannerSeeds"]["assignments"]
    recipes_by_id = {r["id"]: r for r in data["recipes"]}
    prods_by_id   = {p["id"]: p for p in data["products"]}

    day_order = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    slot_order = ["Breakfast","Lunch","Dinner","Snacks"]
    sorted_seeds = sorted(seeds,
        key=lambda a: (day_order.index(a["day"]) if a["day"] in day_order else 99,
                       slot_order.index(a["slot"]) if a["slot"] in slot_order else 99))

    slot_fills = {
        "Breakfast": "FFF7ED",
        "Lunch": "F0FDF4",
        "Dinner": "EFF6FF",
        "Snacks": "FAF5FF"
    }

    for idx, a in enumerate(sorted_seeds):
        r = idx + 3
        item = recipes_by_id.get(a["itemId"]) or prods_by_id.get(a["itemId"], {})
        kcal = item.get("kcalPerServing", item.get("kcal", 0))
        est_kcal = round(kcal * a["servings"], 0) if kcal else "—"
        bg = slot_fills.get(a["slot"], WHITE)
        vals = [a["id"], a["itemId"], a["itemName"],
                a["day"], a["slot"], a["servings"], est_kcal]
        for col, v in enumerate(vals, 1):
            write_cell(ws, r, col, v, bg=bg,
                       align="center" if col in (1,2,4,5,6,7) else "left")

    # -- Shopping list section --
    sl_start = len(seeds) + 5
    ws.cell(row=sl_start, column=1).value = "DERIVED SHOPPING LIST (from seeds above)"
    ws.cell(row=sl_start, column=1).font = bold_font(12, WHITE)
    ws.cell(row=sl_start, column=1).fill = fill(SAGE)
    ws.merge_cells(f"A{sl_start}:E{sl_start}")
    ws.cell(row=sl_start, column=1).alignment = Alignment(horizontal="center",
                                                          vertical="center")
    ws.row_dimensions[sl_start].height = 24

    sl_hdr_row = sl_start + 1
    apply_header_row(ws, ["Ingredient", "Total Qty", "Unit", "Category", "Source"],
                     row=sl_hdr_row, bg=SAGE)

    sl_items = data["plannerSeeds"]["_derivedShoppingList"]["lines"]
    for idx, line in enumerate(sl_items):
        r = sl_hdr_row + 1 + idx
        bg = SAGE_LIGHT if idx % 2 == 0 else WHITE
        vals = [line["ingredient"], line["totalAmount"], line["unit"],
                line["category"], "seed assignments"]
        for col, v in enumerate(vals, 1):
            write_cell(ws, r, col, v, bg=bg,
                       align="center" if col in (2,3) else "left")

    set_col_widths(ws, {1:18, 2:12, 3:24, 4:10, 5:14, 6:14, 7:12})
    ws.row_dimensions[2].height = 26


# ── Sheet 6: Form Inputs ──────────────────────────────────────────────────────
def sheet_form_inputs(wb, data):
    ws = wb.create_sheet("Form Inputs")
    ws.sheet_view.showGridLines = False
    fi = data["formInputs"]

    row = 1
    sections = [
        ("PRODUCT FORM — Valid inputs",      fi["productForm"]["valid"],       TEAL_DARK,  "valid"),
        ("PRODUCT FORM — Edge cases",         fi["productForm"]["edgeCases"],   AMBER,      "edge"),
        ("PRODUCT FORM — Invalid inputs",     fi["productForm"]["invalid"],     "C0392B",   "invalid"),
        ("RECIPE FORM — Valid inputs",        fi["recipeForm"]["valid"],        SAGE,       "valid"),
        ("RECIPE FORM — Invalid inputs",      fi["recipeForm"]["invalid"],      "C0392B",   "invalid"),
        ("PROFILE — Valid macro splits",      fi["profileForm"]["validMacros"], TEAL_MID,   "valid"),
        ("PROFILE — Invalid macro splits",    fi["profileForm"]["invalidMacros"],"C0392B",  "invalid"),
        ("PROFILE — Calorie targets",         fi["profileForm"]["calorieTarget"], TEAL_MID, "edge"),
        ("AUTH — Valid credentials",          fi["authForm"]["validCredentials"], SAGE,     "valid"),
        ("AUTH — Invalid credentials",        fi["authForm"]["invalidCredentials"], "C0392B","invalid"),
    ]

    type_bgs = {"valid": SAGE_LIGHT, "edge": AMBER_LIGHT, "invalid": RED_LIGHT}

    for section_title, items, header_colour, item_type in sections:
        # section heading
        c = ws.cell(row=row, column=1, value=section_title)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = fill(header_colour)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 22
        row += 1

        col_header_row = row
        ws.cell(row=row, column=1, value="Case ID").font   = bold_font(10)
        ws.cell(row=row, column=2, value="Input Data").font = bold_font(10)
        ws.cell(row=row, column=3, value="Expected Result").font = bold_font(10)
        ws.cell(row=row, column=4, value="Notes").font      = bold_font(10)
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = fill(GREY_HEADER)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=col).border = thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

        bg = type_bgs.get(item_type, WHITE)
        for item in items:
            case_id = item.get("_case", "")
            expected = item.get("expectedResult", "")
            notes = item.get("note", item.get("_notes", ""))
            # flatten remaining keys into input summary
            skip_keys = {"_case", "expectedResult", "note", "_notes"}
            input_data = {k: v for k, v in item.items() if k not in skip_keys}
            input_str = "; ".join(f"{k}={v}" for k, v in input_data.items())

            write_cell(ws, row, 1, case_id, bg=bg)
            write_cell(ws, row, 2, input_str, bg=bg, wrap=True)
            write_cell(ws, row, 3, expected, bg=bg, wrap=True)
            write_cell(ws, row, 4, notes, bg=bg, wrap=True)
            ws.row_dimensions[row].height = 45
            row += 1

        row += 1  # blank spacer

    set_col_widths(ws, {1:22, 2:80, 3:60, 4:40})


# ── Sheet 7: Test Coverage Matrix ────────────────────────────────────────────
def sheet_coverage(wb):
    ws = wb.create_sheet("Coverage Matrix")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Req. ID", "Requirement / User Story", "Module",
        "Test Case IDs", "Priority", "Coverage Status"
    ]
    apply_header_row(ws, headers)
    set_col_widths(ws, {1:12, 2:55, 3:22, 4:45, 5:10, 6:18})

    matrix = [
        # AUTH
        ("US-AUTH-01","Registration with unique email + password","Authentication",
         "TC-AUTH-001, TC-AUTH-006, TC-AUTH-008","High","Complete"),
        ("US-AUTH-02","Sign-in with email/password","Authentication",
         "TC-AUTH-002, TC-AUTH-003, TC-AUTH-004","High","Complete"),
        ("US-AUTH-03","Rate-limit on failed sign-in","Authentication",
         "TC-AUTH-005","Medium","Spec only"),
        ("US-AUTH-04","Sign-out clears session","Authentication",
         "TC-AUTH-009, TC-AUTH-010","High","Complete"),
        ("US-AUTH-05","Password reset flow","Authentication",
         "TC-AUTH-011, TC-AUTH-012, TC-AUTH-013","High","Complete"),
        ("US-AUTH-06","Session persists on reload","Authentication",
         "TC-AUTH-014","Medium","Spec only"),
        ("US-AUTH-07","Direct URL redirects unauthenticated user","Authentication",
         "TC-AUTH-015","High","Spec only"),
        # NAV
        ("REQ-NAV","6-item sidebar navigation","Navigation",
         "TC-NAV-001","High","Needs retest"),
        ("REQ-NAV","Active view highlighting","Navigation",
         "TC-NAV-002, TC-NAV-003, TC-NAV-004","Medium","✓ Pass"),
        ("REQ-NAV","Topbar metrics update","Navigation",
         "TC-NAV-005","Medium","✓ Pass"),
        # Products DB
        ("US-PA-001","Browse products with nutrition","Products Database",
         "TC-PRD-001, TC-PRD-009","High","✓ Pass"),
        ("US-PA-002","Filter by category","Products Database",
         "TC-PRD-007, TC-PRD-008, TC-PRD-016","High","✓ Pass"),
        ("US-PA-003","Search by name/attributes","Products Database",
         "TC-PRD-002, TC-PRD-003, TC-PRD-004, TC-PRD-005, TC-PRD-006","High","✓ Pass"),
        ("US-PA-005","Add custom product","Products Database",
         "TC-PRD-017","High","✓ Pass"),
        ("US-PA-006","Edit/delete own products; block on recipe use","Products Database",
         "TC-PRD-018, TC-PRD-023","High","Partial — TC-PRD-023 not tested"),
        ("US-PA-007","This/Next week flags; planner integration","Products Database",
         "TC-PRD-010, TC-PRD-011, TC-PRD-011b, TC-PRD-008b","High","Partial"),
        ("US-PA-008","List/cards view toggle","Products Database",
         "TC-PRD-021, TC-PRD-022","Medium","Not tested"),
        # Products Analyser
        ("REQ-PAN","Spreadsheet nutrition analyser","Products Analyser",
         "TC-PAN-001–010","High","Not implemented"),
        # Recipe Analyser
        ("US-RA-001","Browse recipes with nutrition","Recipe Analyser",
         "TC-RCP-001","High","✓ Pass"),
        ("US-RA-003","Filter by category","Recipe Analyser",
         "TC-RCP-003","High","✓ Pass"),
        ("US-RA-004","Favorites / mine filters","Recipe Analyser",
         "TC-RCP-006, TC-RCP-007","Medium","✓ Pass"),
        ("US-RA-005","Filter by diet","Recipe Analyser",
         "TC-RCP-008","Medium","✓ Pass"),
        ("US-RA-006","Search recipes","Recipe Analyser",
         "TC-RCP-002","High","✓ Pass"),
        ("US-RA-007","Mark favourite","Recipe Analyser",
         "TC-RCP-005","Medium","✓ Pass"),
        ("US-RA-008","This/Next week flags","Recipe Analyser",
         "TC-RCP-013, TC-RCP-016","High","Partial — TC-RCP-016 not tested"),
        ("US-RA-009","Add recipe manually","Recipe Analyser",
         "TC-RCP-011","High","✓ Pass"),
        ("US-RA-010","Import from URL/PDF","Recipe Analyser",
         "TC-RCP-012","Medium","✓ Pass (mock)"),
        ("US-RA-011","Open recipe card","Recipe Analyser",
         "TC-RCP-009, TC-RCP-010","High","✓ Pass"),
        ("US-RA-012","Edit/delete own recipes","Recipe Analyser",
         "TC-RCP-017, TC-RCP-018","High","Not tested"),
        # Dietary Analyser
        ("US-DA-001","View all 12 diets","Dietary Analyser",
         "TC-DIT-003","High","✓ Pass"),
        ("US-DA-002","Diet description + macro split","Dietary Analyser",
         "TC-DIT-002","Medium","✓ Pass"),
        ("US-DA-003","Mark product diet compatibility","Dietary Analyser",
         "TC-DIT-004","Medium","✓ Pass"),
        ("US-DA-004","Mark recipe diet compatibility","Dietary Analyser",
         "TC-DIT-005","Medium","✓ Pass"),
        # Meal Planner
        ("US-MP-001","Week navigation","Meal Planner",
         "TC-PLN-001–004","High","Not tested"),
        ("US-MP-002","Tab switching","Meal Planner",
         "TC-PLN-010, TC-PLN-030, TC-CAL-001","High","Not tested"),
        ("US-MP-003","Weekly summary grid","Meal Planner",
         "TC-PLN-011–015","High","Not tested"),
        ("US-MP-004","Add/remove summary rows","Meal Planner",
         "TC-PLN-018, TC-PLN-019","High","Not tested"),
        ("US-MP-005","Servings/grams toggle","Meal Planner",
         "TC-PLN-016, TC-PLN-017","Medium","Not tested"),
        ("US-MP-006","This-week items auto-populate summary","Meal Planner",
         "TC-PLN-020","High","Not tested"),
        ("US-MP-007","Day cards view","Meal Planner",
         "TC-PLN-030–033","High","Not tested"),
        ("US-MP-009","Adjust servings ±0.5","Meal Planner",
         "TC-PLN-036, TC-PLN-037","Medium","Not tested"),
        ("US-MP-010","Drag and drop","Meal Planner",
         "TC-PLN-038, TC-PLN-039, TC-CAL-009","Low","Not tested"),
        ("US-MP-011","Calendar week view","Meal Planner / Calendar",
         "TC-CAL-001–011","High","Not tested"),
        ("US-MP-012","Calendar month view","Meal Planner / Calendar",
         "TC-CAL-020–030","Medium","Not tested"),
        # Shopping list
        ("US-SL-001","Shopping list as separate nav item","Shopping List",
         "TC-SHP-001","High","Not implemented"),
        ("US-SL-002","Date range selection","Shopping List",
         "TC-SHP-002, TC-SHP-003","High","Not implemented"),
        ("US-SL-003","Plan summary","Shopping List",
         "TC-SHP-004","High","Not implemented"),
        ("US-SL-004","Categorised grocery list","Shopping List",
         "TC-SHP-005–010","High","Not implemented"),
        # Profile
        ("US-PC-001","Change email/password","Personal Cabinet",
         "TC-PRF-007, TC-PRF-013, TC-PRF-014","High","Partial"),
        ("US-PC-002","Language preference","Personal Cabinet",
         "TC-PRF-008","Low","✓ Pass"),
        ("US-PC-003","Unit system","Personal Cabinet",
         "TC-PRF-009","Medium","✓ Pass"),
        ("US-PC-004","Demographic fields","Personal Cabinet",
         "TC-PRF-010","Medium","✓ Pass"),
        ("US-PC-005","Diet preferences + calorie corridor","Personal Cabinet",
         "TC-PRF-001–006","High","✓ Pass"),
        ("US-PC-006","Meal tracking log","Personal Cabinet",
         "TC-MLT-001–005","Medium","✓ Pass (UI present)"),
    ]

    status_fills = {
        "Complete": SAGE_LIGHT,
        "✓ Pass": SAGE_LIGHT,
        "✓ Pass (mock)": SAGE_LIGHT,
        "✓ Pass (UI present)": SAGE_LIGHT,
        "Partial": AMBER_LIGHT,
        "Partial — TC-PRD-023 not tested": AMBER_LIGHT,
        "Partial — TC-RCP-016 not tested": AMBER_LIGHT,
        "Not tested": GREY_HEADER,
        "Needs retest": AMBER_LIGHT,
        "Spec only": TEAL_LIGHT,
        "Not implemented": RED_LIGHT,
    }

    for idx, (req, desc, module, tcs, prio, status) in enumerate(matrix):
        r = idx + 2
        bg = status_fills.get(status, WHITE)
        vals = [req, desc, module, tcs, prio, status]
        for col, v in enumerate(vals, 1):
            row_bg = bg if col == 6 else (TEAL_LIGHT if idx % 2 == 0 else WHITE)
            if col == 6:
                row_bg = bg
            write_cell(ws, r, col, v, bg=row_bg, wrap=(col in (2,4)),
                       align="left")

    ws.row_dimensions[1].height = 26
    for r in range(2, len(matrix) + 2):
        ws.row_dimensions[r].height = 30


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_products(wb, data)
    sheet_recipes(wb, data)
    sheet_diets(wb, data)
    sheet_users(wb, data)
    sheet_planner(wb, data)
    sheet_form_inputs(wb, data)
    sheet_coverage(wb)

    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")
    print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"  Products: {len(data['products'])}")
    print(f"  Recipes: {len(data['recipes'])}")
    print(f"  Diets: {len(data['diets'])}")
    print(f"  Users: {len(data['users'])}")
    print(f"  Planner seeds: {len(data['plannerSeeds']['assignments'])}")


if __name__ == "__main__":
    main()
